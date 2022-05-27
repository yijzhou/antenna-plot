############################################################################################################
# Antenna plots (S11, S21, Efficiency)
# Yijun Zhou
# 9/28/2021
############################################################################################################

import os.path
import csv
import skrf as rf
import matplotlib.pyplot as plt
from openpyxl import load_workbook


def plotS11(datafile, PlotSettingList, bandmark):
    # function to plot S1P/VSWR/Smith from s1p file
    # datafile: list of dictionary, containing file name and label
    # PlotSettingList: list of list
    #   PlotSettingList[0]: list of dictionary for the setting for plot1
    #   PlotSettingList[1]: list of dictionary for the setting for plot2

    plt.figure()    # plot S11
    for dd in datafile:
        snpfile = dd['fname']
        curvelabel = dd['label']
        
        ntwk = rf.Network(snpfile)
        FreqHz = ntwk.f.tolist()
        FreqMHz = [freq/1e6 for freq in FreqHz]   
        S11_dB = list(ntwk.s11.s_db[:,0,0])

        for i in range(len(PlotSettingList[0])):
            plt.subplot(1,len(PlotSettingList[0]),i+1)
            plt.plot(FreqMHz, S11_dB, label=curvelabel)
            plt.legend(loc="lower left")

            plotBandmark(bandmark)
            plt.title(PlotSettingList[0][i]['title'])
            plt.xlim(PlotSettingList[0][i]['x-limit'][0], PlotSettingList[0][i]['x-limit'][1])
            plt.ylim(PlotSettingList[0][i]['y-limit'][0], PlotSettingList[0][i]['y-limit'][1])
            plt.xlabel(PlotSettingList[0][i]['x-label'])
            plt.ylabel(PlotSettingList[0][i]['y-label'])
            plt.grid(1)

    plt.figure()    # plot VSWR
    for dd in datafile:
        snpfile = dd['fname']
        curvelabel = dd['label']

        ntwk = rf.Network(snpfile)
        FreqHz = ntwk.f.tolist()
        FreqMHz = [freq/1e6 for freq in FreqHz]
        VSWR = list(ntwk.s11.s_vswr[:,0,0])
        
        for i in range(len(PlotSettingList[1])):
            plt.subplot(1,len(PlotSettingList[1]),i+1)
            plt.plot(FreqMHz, VSWR, label=curvelabel)
            plt.legend(loc="upper left")

            plotBandmark(bandmark)
            plt.title(PlotSettingList[1][i]['title'])
            plt.xlim(PlotSettingList[1][i]['x-limit'][0],PlotSettingList[1][i]['x-limit'][1])
            plt.ylim(PlotSettingList[1][i]['y-limit'][0],PlotSettingList[1][i]['y-limit'][1])
            plt.xlabel(PlotSettingList[1][i]['x-label'])
            plt.ylabel(PlotSettingList[1][i]['y-label'])
            plt.grid(1)

    """ plt.figure()    # plot Smith Chart
    for dd in datafile:
        snpfile = dd['fname']
        curvelabel = dd['label']

        ntwk = rf.Network(snpfile)
        ntwk.s11.plot_s_smith() """

    plt.show()


def plotS21(datafile, PlotSettingList, bandmark):
    # function to plot S2P from s2p file
    plt.figure()
    for dd in datafile:
        snpfile = dd['fname']
        curvelabel = dd['label']
        
        ntwk = rf.Network(snpfile)
        FreqHz = ntwk.f.tolist()
        FreqMHz = [freq/1e6 for freq in FreqHz]   
        S21_dB = list(ntwk.s21.s_db[:,0,0])

        for i in range(len(PlotSettingList[0])):
            plt.subplot(1,len(PlotSettingList[0]),i+1)
            plt.plot(FreqMHz, S21_dB, label=curvelabel)
            plt.legend(loc="lower left")

            plotBandmark(bandmark)
            plt.title(PlotSettingList[0][i]['title'])
            plt.xlim(PlotSettingList[0][i]['x-limit'][0], PlotSettingList[0][i]['x-limit'][1])
            plt.ylim(PlotSettingList[0][i]['y-limit'][0], PlotSettingList[0][i]['y-limit'][1])
            plt.xlabel(PlotSettingList[0][i]['x-label'])
            plt.ylabel(PlotSettingList[0][i]['y-label'])
            plt.grid(1)

    plt.show()


def plotEff(datafile, PlotSettingList, bandmark):
    # function to plot Efficiency from CSV
    plt.figure()
    for dd in datafile:
        csvfile = dd['fname']
        curvelabel = dd['label']

        FreqMHz, Efficiency = chamber_csv(csvfile)

        for i in range(len(PlotSettingList[0])):
            plt.subplot(1,len(PlotSettingList[0]),i+1)
            plt.plot(FreqMHz, Efficiency, label=curvelabel)
            plt.legend(loc="upper right")

            plotBandmark(bandmark)
            plt.title(PlotSettingList[0][i]['title'])
            plt.xlim(PlotSettingList[0][i]['x-limit'][0], PlotSettingList[0][i]['x-limit'][1])
            plt.ylim(PlotSettingList[0][i]['y-limit'][0], PlotSettingList[0][i]['y-limit'][1])
            plt.xlabel(PlotSettingList[0][i]['x-label'])
            plt.ylabel(PlotSettingList[0][i]['y-label'])
            plt.grid(1)

    plt.show()

def plotBandmark(bandmark):
    # function to plot bandmarks
    if len(bandmark) == 0:
        print ("Missing Bandmark settings...")
        return
    else:
        for i in range(int(len(bandmark)/2)):
            plt.axvspan(bandmark[2*i], bandmark[2*i+1], color='m', alpha=0.05, lw=0)


def find_row (sheet, search_text, match = 1):
    # function to search spreadsheet each row (column[1]) for specific text:
    row = 1

    while True:
        if row > sheet.max_row:
            return 0    # Not found

        if match:
            if search_text == sheet.cell(row,1).value:
                break
            else:
                row = row + 1

        else:
            if search_text in sheet.cell(row,1).value:
                break
            else:
                row = row + 1

    return row  # return row of the found text


def bracket_to_array(string):
    # function to change string ('[]') to float array ([])
    string = string.replace('[','')
    string = string.replace(']','')
    string = string.split(",")
    return [float(ll) for ll in string]


def string_to_array(string):
    # function to change from ['500',...,'6000'] to [500,...,6000]
    string = [ss for ss in string if ss != ""]  # remove empty cell in the last
    return [float(ff) for ff in string]


def read_header (sheet):
    # funtion to read header info in the spreadsheet
    # return value:
    #   datafile: list of dictionary containing file info and label
    #   PlotSettingList: list of dictionary containing plot settings. 
    #       PlotSettingList[1] for the 1st plot, containing a list of dictionary
    #       PlotSettingList[2] for the 2nd plot, containing a list of dictionary
    #   bandmark: float array for plotting band marks

    # data source:
    # search for "Raw data"
    if find_row(sheet, "Raw data"):
        row = find_row(sheet, "Raw data") + 1
    else:
        print("Did not find Raw data in Column 1...")
        return 0, [], []

    datafile = []
    # search 8 rows for raw files
    for i in range(8):
        if sheet.cell(row+i,3).value and sheet.cell(row+i,4).value:
            datafilename = sheet.cell(row+i,3).value + '\\' + sheet.cell(row+i,4).value
            dict = {'fname':datafilename, 'label':sheet.cell(row+i,2).value}
            datafile.append(dict)

    if len(datafile):
        print("%d raw data files to plot..." % len(datafile))
    else:
        print("No raw data file...")
        return 0, [], []

    # plot settings:
    # search for "Plot"
    row = find_row(sheet, "Plot")
    num_plot = 1
    while True:
        if sheet.cell(row,num_plot+2).value:
            num_plot = num_plot + 1
        else:
            break

    # collect plot settings:
    # search for "Layout", find # of subplot:
    row = find_row(sheet, "Layout")
    num_subplot = sheet.cell(row,2).value

    PlotSettingList = [[] for i in range(num_plot)]

    for i in range(num_plot):
        for j in range(num_subplot):
            PlotSettingList[i].append({})


    # start from the first 'Subplot' row
    row = find_row(sheet, "Subplot", 0)

    n = 0
    
    while row <= sheet.max_row:
        if n > num_subplot:
            break

        if sheet.cell(row,1).value: # Not empty row

            if "Subplot" in sheet.cell(row,1).value:
                row = row + 1
                n = n + 1

            else:
                # print("Add row %d to array %d" %(row,n))
                key = sheet.cell(row,1).value
                        
                if "limit" in key:  # change '[]' into list []
                    for i in range(num_plot):
                        PlotSettingList[i][n-1][key] = bracket_to_array(sheet.cell(row,i+2).value)
                else:
                    for i in range(num_plot):
                        PlotSettingList[i][n-1][key] = sheet.cell(row,i+2).value

                row = row + 1
        else:   # if empty row, break the loop
            break
    
    # Find Bandmark setting:
    row = find_row(sheet, "Bandmark")
    if row:
        bands = sheet.cell(row,2).value
        bandmark = []
        while row <= sheet.max_row:
            if sheet.cell(row,1).value: # Not empty row
                if sheet.cell(row,1).value == bands:
                    j = 2
                    while sheet.cell(row,j).value:
                        bandmark.append(sheet.cell(row,j).value)
                        j = j + 1
                    break
                else:
                    row = row + 1
            else:
                break

    else:
        bandmark = []


    return datafile, PlotSettingList, bandmark


def chamber_csv(inFile):
    # function to read CSV from chamber
    # return FreqMHz: frequency list; Efficiency: efficinecy list
    with open(inFile, 'r') as file:
        reader = csv.reader(file)     
        for row in reader:
            if row[0] == "Total" and row[1] == "Frequency  (MHz)" and row[2] != "":
                FreqMHz = string_to_array(row[2:])
                
            if row[0] == "" and row[1] == "Efficiency (dB)":
                Efficiency = string_to_array(row[2:])
                break
                
    return FreqMHz, Efficiency


def main():
    
    #xlsxfilename = 'Auckland_ANT1.xlsx'
    #xlsxfilename = 'Auckland_ANT1 loop on front.xlsx'
    #xlsxfilename = 'Auckland_ANT1 on island2.xlsx'
    xlsxfilename = 'IVAS_GPS.xlsx'
    #xlsxfilename = 'IVAS_ISW.xlsx'
    #xlsxfilename = 'IVAS_WiFi.xlsx'

    if not os.path.isfile(xlsxfilename):
        print ("Input file not exist...")
        return

    # read info from xlx file
    data_xlsx = load_workbook(xlsxfilename, read_only=True)

    if 'S11' in data_xlsx.sheetnames:
        # if there is "S11" spreadsheet
        sheet_S1P = data_xlsx['S11']

        datafile, PlotSettingList, bandmark = read_header(sheet_S1P)
        
        if datafile:
            plotS11(datafile, PlotSettingList, bandmark)

    
    if 'S21' in data_xlsx.sheetnames:
        # if there is "S21" spreadsheet
        sheet_S2P = data_xlsx['S21']

        datafile, PlotSettingList, bandmark = read_header(sheet_S2P)
        
        if datafile:
            plotS21(datafile, PlotSettingList, bandmark)


    if 'Efficiency' in data_xlsx.sheetnames:
        # if there is "Efficiency" spreadsheet
        sheet_Eff = data_xlsx['Efficiency']

        datafile, PlotSettingList, bandmark = read_header(sheet_Eff)

        if datafile:
            plotEff(datafile, PlotSettingList, bandmark)


if __name__ == "__main__":
   main()